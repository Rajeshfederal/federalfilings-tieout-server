"""Builds the Federal Filings side-by-side tie-out workbook (.xlsx).

Takes the row data produced by the AI judgment pass and renders the exact
format of the manually produced workbooks: summary bar with live COUNTIF
formulas, section groupings, currency formatting, red/amber/green statuses,
frozen header, autofilter.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY = "0F2647"; RED_F = "FDE8E8"; RED_T = "B42318"; AMB_F = "FDF0E0"
AMB_T = "B54708"; GRN_F = "E7F4EC"; GRN_T = "067647"; GREY = "EEF2F7"
MM = "MISMATCH"; YL = "YEAR LABEL ERROR"; NS = "NOT IN F/S — CONFIRM"; OK = "MATCH"
CUR = '$#,##0;($#,##0)'


def build_workbook(title, subtitle, rows, out_path):
    """rows: list of dicts with keys:
    section, location, item, mdna_value (num|None), fs_value (num|None),
    fs_source (str), status (MATCH|MISMATCH|YEAR LABEL ERROR|NOT IN F/S — CONFIRM), note (str)
    """
    wb = Workbook(); ws = wb.active; ws.title = "Side-by-Side Tie-Out"
    thin = Border(bottom=Side(style='thin', color='D5DDE7'))
    f_hdr = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    f_sec = Font(name='Arial', bold=True, size=10, color=NAVY)
    f_txt = Font(name='Arial', size=10)

    ws.merge_cells('A1:I1')
    c = ws['A1']; c.value = title
    c.font = Font(name='Arial', bold=True, size=13, color=NAVY)
    ws['A2'] = subtitle
    ws['A2'].font = Font(name='Arial', size=9, color='5B6B7F'); ws.merge_cells('A2:I2')

    hdr = 4
    headers = ["#", "MD&A section", "Location", "Line item / figure", "MD&A value",
               "F/S value", "Difference", "F/S source / computation", "Status", "Reviewer note"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=hdr, column=j, value=h); cell.font = f_hdr
        cell.fill = PatternFill('solid', fgColor=NAVY)
        cell.alignment = Alignment(vertical='center', wrap_text=True)

    r = hdr + 1; last_sec = None; i = 0; first_data = r
    for row in rows:
        sec = row["section"]
        if sec != last_sec:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
            c = ws.cell(row=r, column=1, value=sec)
            c.font = f_sec; c.fill = PatternFill('solid', fgColor=GREY)
            r += 1; last_sec = sec
        i += 1
        st = row["status"]
        fill, tc = (GRN_F, GRN_T)
        if st in (MM, YL): fill, tc = RED_F, RED_T
        elif st == NS: fill, tc = AMB_F, AMB_T
        vals = [i, sec.split(" (")[0], row.get("location", ""), row["item"],
                row.get("mdna_value"), row.get("fs_value"),
                f'=IF(OR(F{r}="",E{r}=""),"",E{r}-F{r})',
                row.get("fs_source", ""), st, row.get("note", "")]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=j, value=v); cell.font = f_txt; cell.border = thin
            if j in (5, 6, 7): cell.number_format = CUR
            if j == 9:
                cell.font = Font(name='Arial', size=10, bold=True, color=tc)
                cell.fill = PatternFill('solid', fgColor=fill)
            if j in (4, 8, 10): cell.alignment = Alignment(wrap_text=True, vertical='top')
        if st in (MM, YL):
            for j in (5, 6, 7): ws.cell(row=r, column=j).fill = PatternFill('solid', fgColor=fill)
        r += 1
    last = r - 1

    ws['A3'] = 'Figures compared:'; ws['B3'] = f'=COUNTA(I{first_data}:I{last})'
    ws['C3'] = 'Mismatches:'
    ws['D3'] = f'=COUNTIF(I{first_data}:I{last},"{MM}")+COUNTIF(I{first_data}:I{last},"{YL}")'
    ws['E3'] = 'Unsupported:'; ws['F3'] = f'=COUNTIF(I{first_data}:I{last},"{NS}")'
    ws['G3'] = 'Match:'; ws['H3'] = f'=COUNTIF(I{first_data}:I{last},"{OK}")'
    for cl in ('A3', 'C3', 'E3', 'G3'): ws[cl].font = Font(name='Arial', size=10, bold=True, color=NAVY)
    ws['B3'].font = Font(name='Arial', size=10, bold=True)
    ws['D3'].font = Font(name='Arial', size=10, bold=True, color=RED_T)
    ws['F3'].font = Font(name='Arial', size=10, bold=True, color=AMB_T)
    ws['H3'].font = Font(name='Arial', size=10, bold=True, color=GRN_T)
    for col, w in {'A': 4, 'B': 26, 'C': 10, 'D': 46, 'E': 14, 'F': 14, 'G': 12,
                   'H': 42, 'I': 22, 'J': 46}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f'A{hdr + 1}'
    ws.auto_filter.ref = f"A{hdr}:J{last}"
    wb.save(out_path)
    return out_path
